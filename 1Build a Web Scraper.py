# Level 3 :Task 1
# Task: Build a Web Scraper
'''
 Develop a web scraper that extracts specific data from websites using libraries like BeautifulSoup or Scrapy.
 This task will improve their knowledge of web scraping techniques and handling HTML/XML data.
'''

import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def scrape_webpage(url):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/114.0.0.0 Safari/537.36"
        }
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        wb = Workbook()
        ws_info = wb.active
        ws_info.title = "Page Info"
        ws_headings = wb.create_sheet("Headings")
        ws_paragraphs = wb.create_sheet("Paragraphs")
        ws_images = wb.create_sheet("Images")
        ws_links = wb.create_sheet("Links")

        page_title = soup.title.string.strip() if soup.title else 'No Title'
        ws_info.append(["Page Title", page_title])
        print(f"Page Title: {page_title}")

        ws_headings.append(["Tag", "Text"])
        for tag in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
            for h in soup.find_all(tag):
                text = h.get_text(strip=True)
                if text:
                    ws_headings.append([tag.upper(), text])

        ws_paragraphs.append(["Index", "Text"])
        paragraphs = soup.find_all('p')
        for i, p in enumerate(paragraphs, 1):
            ws_paragraphs.append([i, p.get_text(strip=True)])

        ws_images.append(["Index", "Image URL"])
        images = soup.find_all('img')
        for i, img in enumerate(images, 1):
            src = img.get('src')
            if src:
                ws_images.append([i, urljoin(url, src)])

        ws_links.append(["Index", "Link Text", "URL"])
        links = soup.find_all('a', href=True)
        for i, link in enumerate(links, 1):
            text = link.get_text(strip=True) or "No Text"
            href = urljoin(url, link['href'])
            ws_links.append([i, text, href])

        for sheet in wb.worksheets:
            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        filename = "scraped_data.xlsx"
        wb.save(filename)
        print(f"Data scraped and saved to '{filename}' successfully.")

    except requests.exceptions.RequestException as e:
        print(f"Request failed: {e}")
    except Exception as e:
        print(f"Unexpected error: {e}")

if __name__ == "__main__":
    url = input("Enter any webpage URL: ").strip()
    scrape_webpage(url)